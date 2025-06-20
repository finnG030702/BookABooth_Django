from django.contrib import messages
from django.contrib.auth import get_user_model, logout, authenticate
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.contrib.auth.tokens import default_token_generator
from django.shortcuts import get_object_or_404, render, redirect
from django.urls import reverse_lazy, reverse
from django.utils import timezone
from django.utils.safestring import mark_safe
from django.utils.http import urlsafe_base64_encode, urlsafe_base64_decode
from django.utils.encoding import force_bytes
from django.views import View
from django.views.decorators.http import require_POST
from django.http import HttpResponse, HttpResponseBadRequest, HttpResponseForbidden, HttpResponseRedirect, JsonResponse
from django.core.mail import send_mail
from django.db.models import Prefetch, Count, Q, Exists, OuterRef, Subquery
from types import SimpleNamespace
from decimal import Decimal
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from django.views.generic import TemplateView, CreateView, ListView, DetailView, UpdateView, DeleteView
from django.contrib.auth.views import LoginView, PasswordChangeView, PasswordResetView, PasswordResetCompleteView, PasswordResetDoneView, PasswordResetConfirmView
from .forms import CompanyForm, CustomPasswordResetForm, CustomUserChangeForm, CustomUserCreationForm, CustomUserLoginForm, CustomPasswordChangeForm, CustomPasswordResetConfirmForm, LocationForm, BoothForm, ServicePackageForm, CustomCompanyChangeForm
from .models import Company, System, Location, Booth, ServicePackage, Booking, SystemConfiguration, TermsUpdateLog

User = get_user_model()


# Create your views here.

@login_required
@require_POST
def accept_privacy_policy(request):
    user = request.user
    user.privacy_policy_accepted = True
    user.save()
    return redirect("home")

@login_required
@require_POST
def add_to_waiting_list(request):
    user = request.user
    if hasattr(user, "company"):
        company = user.company
        company.waiting_list = True
        company.save()
        messages.success(request, "Sie wurden der Warteliste hinzugefügt.")
        return JsonResponse({"success": True})
    return JsonResponse({"success": False, "error": "Kein Unternehmen gefunden"}, status=400)

class HomeView(TemplateView):
    """
    View for the homepage. Contains data for User-Checklist and for Admin-Panel.
    """
    template_name = "home.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        system = System.objects.get(id=1)
        context["system"] = system or System(enabled=False) # Check if booking is enabled / disabled by default

        user = self.request.user

        if user.is_authenticated and not user.is_staff:
            company = getattr(user, "company", None)


            checklist = {
                "registered": True, # Account is created
                "verified": user.is_active, # E-Mail verification completed
                "phone": bool(user.phone),
                "address": bool(company and company.billing_address),
                "logo": bool(company and company.logo),
                "description": bool(company and company.description),
                "waitingList": bool(company and company.waiting_list),
                "privacyPolicy": bool(user.privacy_policy_accepted),
                "allBoothsOccupied": not Booth.objects.filter(available=True).exists(),
            }

            booking_status = None
            if company:
                latest_booking = company.bookings.order_by('-received').first()
                if latest_booking:
                    booking_status = latest_booking.status.upper()
            
            checklist["bookingStatus"] = booking_status
            context["checklist"] = SimpleNamespace(**checklist)

        elif user.is_authenticated and user.is_staff:
            locations = Location.objects.annotate(
                total_booths=Count('booths'),
                booked_booths=Count('booths__bookings', filter=Q(booths__bookings__status='confirmed'))
            )

            location_data = []
            for loc in locations:
                total = loc.total_booths or 0
                booked = loc.booked_booths or 0
                percent = (booked / total * 100) if total > 0 else 0
                location_data.append({
                    "name": loc.location,
                    "total_booths": total,
                    "booked_booths": booked,
                    "percent_booked": percent,
                })

            companies = Company.objects.prefetch_related(
                'employees',
                Prefetch(
                    'bookings',
                    queryset=Booking.objects.select_related('booth__location')
                )
            ).all()

            company_data = []
            for company in companies:
                company_data.append({
                    "name": company.name,
                    "billing_address": bool(company.billing_address),
                    "logo": bool(company.logo),
                    "description": bool(company.description),
                    "phone": bool(company.employees.first().phone),
                    "bookings": company.bookings.all(),
                })
            
            user_with_email = User.objects.filter(
                Q(email__isnull=False) & ~Q(email='') & Q(is_superuser=False)
            )
            mail_recipients = user_with_email.values_list('email', flat=True)
            mailto_all = "mailto:?bcc=" + ";".join(mail_recipients)

            incomplete_users = user_with_email.filter(
                Q(phone__isnull=True) | Q(phone__exact='') |
                Q(company__billing_address__isnull=True) | Q(company__billing_address__exact='') |
                Q(company__logo__isnull=True) | 
                Q(company__description__isnull=True) | Q(company__description__exact='')
            )
            incomplete_recipients = incomplete_users.values_list('email', flat=True)
            mailto_incomplete = "mailto:?bcc=" + ";".join(incomplete_recipients)

            context["admin_location_data"] = location_data
            context["admin_company_data"] = company_data
            context["total_mail"] = user_with_email.count()
            context["incomplete_mail"] = incomplete_users.count()
            context["mailto_all"] = mailto_all
            context["mailto_incomplete"] = mailto_incomplete

        return context

class SignUpView(CreateView):
    """
    View for registration of a new user. The company comes from a different model.
    """
    form_class = CustomUserCreationForm
    success_url = reverse_lazy("login")
    template_name = "registration/signup.html"

    def form_valid(self, form):
        company_name = form.cleaned_data["company_name"]
        company, _ = Company.objects.get_or_create(name=company_name) # Gibt Company und ein Boolean zurück. _ weil created (bool) egal ist.

        user = form.save(commit=False)
        user.company = company
        user.is_active = False
        user.privacy_policy_accepted = form.cleaned_data['privacy_policy_accepted']
        user.save()

        self.object = user
        self.send_verification_email(user)

        messages.success(self.request, "Registrierung erfolgreich! Bitte überprüfen Sie Ihre E-Mails zur Verifizierung.")
        return HttpResponseRedirect(self.get_success_url())
    
    def send_verification_email(self, user):
        token = default_token_generator.make_token(user)
        uid = urlsafe_base64_encode(force_bytes(user.pk))
        verification_link = self.request.build_absolute_uri(
            reverse("verify_email", kwargs={"uidb64": uid, "token": token})
        )
        send_mail(
            subject="Django BookABooth - Bitte bestätigen Sie Ihre E-Mail",
            message=(
                f"Moin {user.username}, \n\n"
                f"bitte klicken Sie auf den folgenden Link, um Ihr Nutzerkonto zu aktivieren:\n"
                f"{verification_link}\n\n"
                f"Falls Sie sich nicht registriert haben, sollten Sie dies vielleicht tun."
            ),
            from_email="noreply@bookabooth.de",
            recipient_list=[user.email],
        )
    
def verify_email(request, uidb64, token):
    """
    Verification for Mail sent in Signup. Might work.
    """
    try:
        uid = urlsafe_base64_decode(uidb64).decode()
        user = User.objects.get(pk=uid)
    except (TypeError, ValueError, OverflowError, User.DoesNotExist):
        user = None
    
    token_valid = user and default_token_generator.check_token(user, token)
    if token_valid:
        user.is_active = True
        user.save()
        return render(request, "registration/verify_success.html")
    else:
        return render(request, "registration/verify_failed.html")

class LoginView(LoginView):
    """
    View for the Login of a user. If logged in, reverts back to home.html
    """
    form_class = CustomUserLoginForm
    success_url = reverse_lazy("login")
    template_name = "registration/login.html"

class CustomPasswordChangeView(PasswordChangeView):
    """
    View for the Password Change. If changed, directs to password-change/done.
    """
    form_class = CustomPasswordChangeForm
    template_name = 'registration/password_change_form.html'

class PasswortResetView(PasswordResetView):
    template_name = "registration/password_reset.html"
    form_class = CustomPasswordResetForm

class PasswordResetDoneView(PasswordResetDoneView):
    template_name = "registration/password_reset_done.html"

class PasswordResetConfirmView(PasswordResetConfirmView):
    template_name = "registration/password_reset_confirm.html"
    form_class = CustomPasswordResetConfirmForm

class PasswordResetCompleteView(PasswordResetCompleteView):
    template_name = "registration/password_reset_complete.html"

class SystemToggleView(LoginRequiredMixin, UserPassesTestMixin, View):
    """
    View for the SystemToggle.html. Changes the enabled-Attribute of system in its response.
    """
    template_name = "adminMenu/system/system_toggle.html"

    def test_func(self):
        return self.request.user.is_superuser

    def get(self, request):
        system = System.objects.first()
        return render(request, self.template_name, {"system": system})
    
    def post(self, request):
        system = System.objects.first()
        enabled = request.POST.get("enabled") == "on"
        system.enabled = enabled
        system.save()
        return render(request, "adminMenu/system/_system_status.html", {"system": system})
    
class LocationListView(LoginRequiredMixin, UserPassesTestMixin, ListView):
    model = Location
    template_name = "adminMenu/location/location_list.html"
    context_object_name = 'locations'

    def test_func(self):
        return self.request.user.is_superuser

    def location_table_partial(request):
        locations = Location.objects.all()
        return render(request, "adminMenu/location/location_table_body.html", {"locations": locations})

class LocationCreateView(LoginRequiredMixin, UserPassesTestMixin, CreateView):
    model = Location
    form_class = LocationForm
    template_name = "adminMenu/location/location_form.html"
    success_url = reverse_lazy("location_list")

    def test_func(self):
        return self.request.user.is_superuser

class LocationDetailView(LoginRequiredMixin, UserPassesTestMixin, DetailView):
    model = Location
    template_name = "adminMenu/location/location_detail.html"

    def test_func(self):
        return self.request.user.is_superuser

class LocationUpdateView(LoginRequiredMixin, UserPassesTestMixin, UpdateView):
    model = Location
    form_class = LocationForm
    template_name = "adminMenu/location/location_form.html"
    success_url = reverse_lazy("location_list")

    def test_func(self):
        return self.request.user.is_superuser

class LocationDeleteView(LoginRequiredMixin, UserPassesTestMixin, DeleteView):
    model = Location
    success_url = reverse_lazy("location_list")

    def test_func(self):
        return self.request.user.is_superuser

    def delete(self, request, *args, **kwargs):
        return super().delete(request, *args, **kwargs)
    
class BoothListView(LoginRequiredMixin, UserPassesTestMixin, ListView):
    model = Booth
    template_name = "adminMenu/booth/booth_list.html"
    context_object_name = 'booths'

    def test_func(self):
        return self.request.user.is_superuser

    def booth_table_partial(request):
        booths = Booth.objects.all()
        return render(request, "adminMenu/booth/booth_table_body.html", {"booths": booths})
    
class BoothCreateView(LoginRequiredMixin, UserPassesTestMixin, CreateView):
    model = Booth
    form_class = BoothForm
    template_name = "adminMenu/booth/booth_form.html"
    success_url = reverse_lazy("booth_list")

    def test_func(self):
        return self.request.user.is_superuser

class BoothDetailView(LoginRequiredMixin, UserPassesTestMixin, DetailView):
    model = Booth
    template_name = "adminMenu/booth/booth_detail.html"

    def test_func(self):
        return self.request.user.is_superuser

class BoothUpdateView(LoginRequiredMixin, UserPassesTestMixin, UpdateView):
    model = Booth
    form_class = BoothForm
    template_name = "adminMenu/booth/booth_form.html"
    success_url = reverse_lazy("booth_list")

    def test_func(self):
        return self.request.user.is_superuser

class BoothDeleteView(LoginRequiredMixin, UserPassesTestMixin, DeleteView):
    model = Booth
    success_url = reverse_lazy("booth_list")

    def test_func(self):
        return self.request.user.is_superuser

    def delete(self, request, *args, **kwargs):
        return super().delete(request, *args, **kwargs)
    
class ServicepackageListView(LoginRequiredMixin, UserPassesTestMixin, ListView):
    model = ServicePackage
    template_name = "adminMenu/servicepackage/servicepackage_list.html"
    context_object_name = 'servicepackages'

    def test_func(self):
        return self.request.user.is_superuser

    def servicepackage_table_partial(request):
        servicepackages = ServicePackage.objects.all()
        return render(request, "adminMenu/servicepackage/servicepackage_table_body.html", {"servicepackages": servicepackages})
    
class ServicepackageCreateView(LoginRequiredMixin, UserPassesTestMixin, CreateView):
    model = ServicePackage
    form_class = ServicePackageForm
    template_name = "adminMenu/servicepackage/servicepackage_form.html"
    success_url = reverse_lazy("servicepackage_list")

    def test_func(self):
        return self.request.user.is_superuser

class ServicepackageDetailView(LoginRequiredMixin, UserPassesTestMixin, DetailView):
    model = ServicePackage
    template_name = "adminMenu/servicepackage/servicepackage_detail.html"

    def test_func(self):
        return self.request.user.is_superuser

class ServicepackageUpdateView(LoginRequiredMixin, UserPassesTestMixin, UpdateView):
    model = ServicePackage
    form_class = ServicePackageForm
    template_name = "adminMenu/servicepackage/servicepackage_form.html"
    success_url = reverse_lazy("servicepackage_list")

    def test_func(self):
        return self.request.user.is_superuser

class ServicepackageDeleteView(LoginRequiredMixin, UserPassesTestMixin, DeleteView):
    model = ServicePackage
    success_url = reverse_lazy("servicepackage_list")

    def test_func(self):
        return self.request.user.is_superuser

    def delete(self, request, *args, **kwargs):
        return super().delete(request, *args, **kwargs)
    
class BookingListView(LoginRequiredMixin, UserPassesTestMixin, ListView):
    model = Booking
    template_name = "adminMenu/booking/booking_list.html"
    context_object_name = "bookings"

    def test_func(self):
        return self.request.user.is_superuser

    def booking_table_partial(request):
        bookings = Booking.objects.all()
        return render(request, "adminMenu/booking/booking_table_body.html", {'bookings': bookings})
    
class BookingDetailView(LoginRequiredMixin, UserPassesTestMixin, DetailView):
    model = Booking
    template_name = "adminMenu/booking/booking_detail.html"

    def test_func(self):
        return self.request.user.is_superuser

class BookingCancelView(LoginRequiredMixin, UserPassesTestMixin, DeleteView):
    def post(self, request, *args, **kwargs):
        booking = get_object_or_404(Booking, pk=kwargs['pk'])

        try:
            booking.cancel(canceled_by_admin=True)
            messages.success(request, "Buchung wurde erfolgreich storniert.")
        except ValueError as e:
            messages.error(request, str(e))

        return redirect("booking_list")

    def test_func(self):
        return self.request.user.is_superuser

class CompanyDetailView(LoginRequiredMixin, UserPassesTestMixin, DetailView):
    model = Company
    template_name = "adminMenu/company/company_detail.html"

    def test_func(self):
        return self.request.user.is_superuser

class CompanyUpdateView(LoginRequiredMixin, UserPassesTestMixin, UpdateView):
    model = Company
    form_class = CompanyForm
    template_name = "adminMenu/company/company_form.html"

    def test_func(self):
        return self.request.user.is_superuser

    def get_success_url(self):
        return reverse_lazy('company_detail', kwargs={'pk': self.object.pk})
    
class WaitingListView(LoginRequiredMixin, UserPassesTestMixin, ListView):
    model = Company
    template_name = "adminMenu/waitingList/waitingList.html"
    context_object_name = "companies"

    def test_func(self):
        return self.request.user.is_superuser

    def get_queryset(self):
        return Company.objects.order_by('-waiting_list', 'id')

    def waitinglist_table_partial(request):
        companies = Company.objects.order_by('-waiting_list', 'id')
        return render(request, "adminMenu/waitingList/waitinglist_table_body.html", {'companies': companies})
    
    def add_to_waitinglist(request, company_id):
        company = get_object_or_404(Company, id=company_id)
        company.waiting_list = True
        company.save()
        companies = Company.objects.order_by('-waiting_list', 'id')
        return render(request, "adminMenu/waitingList/waitinglist_table_body.html", {"companies": companies})
    
    def remove_from_waitinglist(request, company_id):
        company = get_object_or_404(Company, id=company_id)
        company.waiting_list = False
        company.save()
        companies = Company.objects.order_by('-waiting_list', 'id')
        return render(request, "adminMenu/waitingList/waitinglist_table_body.html", {"companies": companies})
    
    def notify_waitinglist_companies(request):
        companies = Company.objects.filter(waiting_list=True)

        for company in companies:
            users = company.employees.all()
            for user in users:
                message = (
                    f"Moin {user.username},\n\n"
                    "Sie stehen beim diesjährigen Jade Karrieretag auf der Warteliste. "
                    "Soeben sind wieder Stände zur Buchung verfügbar geworden.\n"
                    "Eine Reservierung kann über das Buchungstool durchgeführt werden.\n\n"
                    "Sollten Sie diese Benachrichtigungen nicht mehr empfangen wollen, "
                    "können Sie sich im Buchungstool von der Warteliste entfernen.\n\n"
                    "Mit freundlichen Grüßen\n"
                    "Jade Hochschule"
                )
                send_mail(
                    subject="Jade Karrieretag Warteliste",
                    message=message,
                    from_email="noreply@bookabooth.de",
                    recipient_list=[user.email],
                    fail_silently=False,
                )

@login_required
@require_POST
def send_waitinglist_email(request):
    if not request.user.is_superuser:
        return JsonResponse({"success": False, "message": "Nicht autorisiert"}, status=403)
    
    WaitingListView.notify_waitinglist_companies(request)
    messages.success(request, "Unternehmen auf Warteliste wurden benachrichtigt.")
    
    response = HttpResponse(status=204)
    response['HX-Redirect'] = reverse("waitingList")
    return response

class TermsResetView(LoginRequiredMixin, UserPassesTestMixin, View):
    template_name = "adminMenu/privacyPolicy/privacyPolicy.html"

    def test_func(self):
        return self.request.user.is_superuser

    def get(self, request):
        update_logs = TermsUpdateLog.objects.order_by('-updated_at')
        return render(request, self.template_name, {"logs": update_logs})

    def post(self, request):
        User.objects.update(privacy_policy_accepted=False)
        TermsUpdateLog.objects.create()
        return redirect("privacyPolicy")
    
class ProfileView(LoginRequiredMixin, View):

    def _get_company(self, user):
        """
        Gets the company to the corresponding user
        """
        return getattr(user, 'company', None)
    
    def _has_booking(self, company, status):
        if not company:
            return False
        return Booking.objects.filter(company=company, status__in=status).exists()
    
    def _is_only_one_admin(self):
        return User.objects.filter(is_superuser=True, is_active=True).count() == 1
    
    def _handle_redirect(self, request, url_name):
        """
        Returns HTMX-Redirect for HTMX-Requests. Otherwise, default Django-Redirect
        """
        url = reverse(url_name)
        if request.htmx:
            response = HttpResponse()
            response['HX-Redirect'] = url
            return response
        else:
            return redirect(url)
        
    def _calculate_cancellation_fee(self, booking, system_config):
        """
        Calculates the fee for canceling a booking.
        """
        if not booking or not system_config:
            return Decimal("0.00")
        
        multiplier = Decimal("1.0")

        if timezone.now().date() <= system_config.cancellation_reimbursement_until:
            multiplier = Decimal(100 - system_config.cancellation_reimbursement) / Decimal(100)

        return (booking.price * multiplier).quantize(Decimal("0.01"))
    
    def dispatch(self, request, *args, **kwargs):
        action = kwargs.get('action')
        if action == 'modal_cancel_booking' and request.method == 'GET':
            return self.get_modal_cancel_booking(request)
        elif action == 'modal_delete_account' and request.method == 'GET':
            return self.get_modal_delete_account(request)
        return super().dispatch(request, *args, **kwargs)

    def get(self, request):
        user = request.user
        company = self._get_company(user)

        context = {
            "user_data_form": CustomUserChangeForm(instance=user),
            "company_data_form": CustomCompanyChangeForm(instance=company) if company else None,
            "has_confirmed_booking": self._has_booking(company, ['confirmed']),
            "has_canceled_booking": self._has_booking(company, ['canceled']),
            "is_admin": user.is_superuser,
            "only_one_admin": self._is_only_one_admin() if user.is_superuser else False,
        }
        return render(request, "settings/profile.html", context)
    
    def post(self, request, *args, **kwargs):
        # Unterschied, welcher Button im Profil genutzt wird
        if "cancel_booking" in request.POST:
            return self.cancel_booking(request)
        elif "leave_waiting_list" in request.POST:
            return self.leave_waiting_list(request)
        elif "delete_account" in request.POST:
            return self.delete_account(request)
        else:
            return self.save(request)
    
    def save(self, request):
        user = request.user
        company = self._get_company(user)

        user_data = CustomUserChangeForm(request.POST, instance=user)
        company_data = CustomCompanyChangeForm(request.POST, request.FILES, instance=user.company) if company else None

        if user_data.is_valid() and (company_data is None or company_data.is_valid()):
            user_data.save()
            messages.success(request, mark_safe(
                f"<strong>Die Einstellungen wurden gespeichert.</strong> Weiter geht es auf der "
                f"<a href='{reverse('home')}' class='text-blue-600 hover:underline'>Startseite</a>."
                )
            )
            if company_data:
                company_data.save()
            return self._handle_redirect(request, 'profile')
        
        context = {
            "user_data_form": user_data,
            "company_data_form": company_data
        }
        return render(request, "settings/profile.html", context)
    
    def cancel_booking(self, request, *args, **kwargs):
        user = request.user
        company = self._get_company(user)

        if not company:
            messages.error(request, "Kein Unternehmen gefunden")
            return self._handle_redirect(request, 'profile')
        
        try:
            booking = Booking.objects.get(company=company, status='confirmed')
        except Booking.DoesNotExist:
            messages.error(request, "Keine bestätigte Buchung zum Stornieren gefunden.")
            return self._handle_redirect(request, 'profile')

        try:
            booking.cancel()
        except ValueError as e:
            messages.error(request, str(e))
            return self._handle_redirect(request, 'profile')

        messages.success(request, "Ihre Buchung wurde storniert.")
        return self._handle_redirect(request, 'profile')
    
    def leave_waiting_list(self, request, *args, **kwargs):
        user = request.user
        company = self._get_company(user)

        if not company:
            messages.error(request, "Kein Unternehmen gefunden.")
            return self._handle_redirect(request, 'profile')
        
        if not company.waiting_list:
            messages.info(request, "Sie stehen nicht auf der Warteliste.")
            return self._handle_redirect(request, 'profile')
        
        company.waiting_list = False
        company.save()
        messages.success(request, "Sie wurden von der Warteliste entfernt.")
        return self._handle_redirect(request, 'profile')
    
    def delete_account(self, request, *args, **kwargs):
        user = request.user
        company = self._get_company(user)

        # Passwort auslesen und prüfen
        password = request.POST.get("password")
        if not password or not authenticate(username=user.username, password=password):
            messages.error(request, "Passwort ist falsch oder fehlt.")
            return self._handle_redirect(request, 'profile')

        # Prüfen, ob der User ein Admin ist und ob mehrere Admins im System sind
        if user.is_superuser:
            if self._is_only_one_admin():
                messages.error(request, "Der einzige Admin kann nicht gelöscht werden.")
                return self._handle_redirect(request, 'profile')
            
        # Prüfen, ob Booking existiert
        has_booking = self._has_booking(company, ['confirmed', 'canceled'])
        if has_booking:
            messages.error(request, "Konto kann nicht gelöscht werden, solange eine Buchung besteht.")
            return self._handle_redirect(request, 'profile')
        
        # Wenn keine Buchung existiert, User löschen
        if company:
            company.delete()
        user.delete()
        logout(request)
        messages.success(request, "Ihr Konto wurde erfolgreich gelöscht.")
        return self._handle_redirect(request, 'home')
    
    def get_modal_cancel_booking(self, request, *args, **kwargs):
        user = request.user
        company = self._get_company(user)
        system_config = SystemConfiguration.objects.first()

        try:
            booking = Booking.objects.get(company=company, status='confirmed')
        except Booking.DoesNotExist:
            return HttpResponse("Keine bestätigte Buchung gefunden", status=404)

        cancellation_fee = self._calculate_cancellation_fee(booking, system_config)
        context = {
            "system": system_config,
            "cancellation_fee": cancellation_fee,
            "booking": booking,
        }

        if request.htmx:
            return render(request, "settings/cancel_booking.html", context)
        return HttpResponseBadRequest("Invalid request")
    
    def get_modal_delete_account(self, request, *args, **kwargs):
        if request.htmx:
            return render(request, "settings/delete_account.html")
        return HttpResponseBadRequest("Invalid request")
    
class BookABoothView(LoginRequiredMixin, TemplateView):
    template_name = "bookabooth.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        user_company = None

        selected_location_id = self.request.GET.get('location')
        locations = Location.objects.all()

        if selected_location_id:
            booths = Booth.objects.filter(location_id=selected_location_id)
        else:
            booths = Booth.objects.all()

        for booth in booths:
            # Calculate total price for the booth
            booth.total_price = sum(package.price for package in booth.service_package.all())

            # Display the Company that booked the booth, checks for confirmed booking and exhibitor_list
            booth.company_name = None
            booking = Booking.objects.filter(booth=booth, status='confirmed').order_by('-received').first()
            if booking and booking.company and booking.company.exhibitor_list:
                booth.company_name = booking.company.name

        if self.request.user.is_authenticated and hasattr(self.request.user, 'company'):
            user_company = self.request.user.company

        # Checks if the User already has a confirmed booking. If the booking is canceled, a new one can be created. Also saves the booth.
        has_booking = False
        booking = None
        booth = None
        if user_company:
            booking = Booking.objects.filter(company=user_company, status='confirmed').order_by("-received").first()
            has_booking = booking is not None
            if booking:
                booth = booking.booth

        # Displays the Site_Plan for a location, if one is provided.
        location = None
        location_map = None
        if not has_booking and selected_location_id:
            try:
                location = Location.objects.get(id=selected_location_id)
                location_map = location.site_plan.url if location.site_plan else None
            except Location.DoesNotExist:
                location_map = None

        if user_company:
            checklist_complete = all([
                user_company.billing_address,
                user_company.logo,
                self.request.user.phone,
                user_company.description
            ])
        else:
            checklist_complete = True

        context.update({
            'has_booking': has_booking,
            'booking': booking,
            'booth': booth,
            'locations': locations,
            'booths': booths,
            'selected_location': int(selected_location_id) if selected_location_id else None,
            'location_map': location_map,
            'location': location if selected_location_id else None,
            'checklist_complete': checklist_complete,
        })

        return context

def booking_modal(request, booth_id):
    """
    Is called when BookingModal is opened.
    Checks that the User is logged in and hasn't already booked a booth.
    Sets the booking status to 'blocked' and makes the selected booth unavailable.
    """
    booth = get_object_or_404(Booth, id=booth_id)

    # Check if User is logged in and has a company
    if not request.user.is_authenticated or not hasattr(request.user, 'company'):
        return HttpResponseForbidden("Bitte melden Sie sich an, um einen Stand zu buchen.")
    
    user = request.user
    company = user.company

    checklist_complete = all([
        company.billing_address,
        company.logo,
        user.phone,
        company.description
    ])

    # Check if the profile is complete
    if not checklist_complete:
        return HttpResponseForbidden("Bitte vervollständigen Sie Ihr Profil, bevor Sie einen Stand buchen.")
    
    # Check if privacy policy was accepted
    if not user.privacy_policy_accepted:
        return HttpResponseForbidden("Bitte akzeptieren Sie die Nutzungsbedingungen, bevor Sie einen Stand buchen.")

    # Check if user has a confirmed booking
    has_other_booking = Booking.objects.filter(
        company=company,
        status='confirmed'
    ).exclude(booth=booth).exists()

    if has_other_booking:
        return HttpResponseForbidden("Sie haben bereits einen anderen Stand gebucht.")
    
    # Check if user has a blocked booking
    has_blocked_booking = Booking.objects.filter(
        company=company,
        status='blocked'
    ).exclude(booth=booth).exists()

    if has_blocked_booking:
        return HttpResponseForbidden("Sie haben bereits eine Buchung in Bearbeitung. Warten Sie ein paar Minuten und versuchen Sie es erneut.")
    
    total_price = sum(package.price for package in booth.service_package.all())

    # TODO: Wird das Booking auch erstellt, wenn das Modal sofort wieder geschlossen wird?
    # TODO: Außerdem vielleicht redirect wenn auf Abbrechen?
    booking = Booking.objects.create(
        booth=booth,
        company=company,
        received=timezone.now(),
        status='blocked',
        price=total_price
    )

    system_config = SystemConfiguration.objects.first()

    # Set booth to 'unavailable' to prevent other bookings
    booth.available = False
    booth.save()

    return render(request, "components/booking_modal.html", {
        'booth': booth,
        'booking': booking,
        'system': system_config
    })

@require_POST
@login_required
def confirm_booking(request, booth_id):
    """
    Called when the user confirms the booking.
    Sets booking to 'confirmed' and saves the current time.
    Redirect is handled with HTMX.
    """
    booth = get_object_or_404(Booth, id=booth_id)
    company = request.user.company
    booking = Booking.objects.filter(booth=booth, company=company, status='blocked').first()

    if not booking:
        return JsonResponse({'error': 'Keine gültige Buchung gefunden.'}, status=404)

    booking.status = 'confirmed'
    booking.confirmed = timezone.now()
    booking.save()

    messages.success(request, "Ihre Buchung wurde erfolgreich bestätigt!")

    response = HttpResponse()
    response['HX-Redirect'] = reverse('home')
    return response

def exhibitor_info(request):
    return render(request, "components/ausstellerinfo.html")

@login_required
@user_passes_test(lambda u: u.is_superuser)
def export_to_excel(request):
    response = HttpResponse(content_type='application/ms-excel')
    filename = f"Buchungen-{timezone.now().strftime('%Y-%m-%d')}.xlsx"
    response["Content-Disposition"] = f"attachment; filename=\"{filename}\""

    wb = Workbook()
    ws = wb.active
    ws.title = "Buchungen"

    headers = ["Firmenname", "Ansprechpartner", "Rechnungsadresse", "Bemerkung", "Standnummer", "Preis", "Stornierung", "Rechnungsnummer", "Innenauftrag", "Kundennummer"]
    ws.append(headers)

    bookings = Booking.objects.select_related(
        'booth__location',
        'company'
    ).all()

    for b in bookings:
        # Get every employee of a company
        employees = b.company.employees.all() if b.company else []

        # bring the user into lastname, firstname (phone)-format
        if employees:
            user = employees[0]
            parts = []
            if user.last_name:
                parts.append(user.last_name)
            if user.first_name:
                parts.append(user.first_name)
            ansprechpartner = ", ".join(parts)
            if user.phone:
                ansprechpartner += f" ({user.phone})"
        else:
            ansprechpartner = ""

        ws.append([
            b.company.name if b.company else "",
            ansprechpartner,
            b.company.billing_address if b.company and hasattr(b.company, "billing_address") else "",
            b.company.comment if b.company and hasattr(b.company, "comment") else "",
            f"{b.booth.location.location} - {b.booth.title}" if b.booth and b.booth.location else "",
            b.price if hasattr(b, "price") else None,
            b.cancellationfee if hasattr(b, "cancellationfee") else None
        ])

        last_row = ws.max_row
        for col_letter in ['F', 'G']:
            cell = ws[f"{col_letter}{last_row}"]
            if isinstance(cell.value, (int, float)):
                cell.number_format = '#.##0,00'

    for col in ws.columns:
        max_length = 0
        column = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    wb.save(response)
    return response

def exhibitorlistView(request):
    confirmed_bookings = Booking.objects.filter(company=OuterRef('pk'), status='confirmed')
    booth_subquery = Booking.objects.filter(company=OuterRef('pk'), status='confirmed').values('booth__title')[:1]
    location_subquery = Booking.objects.filter(company=OuterRef('pk'), status='confirmed').values('booth__location__location')[:1]

    companies = Company.objects.filter(
        exhibitor_list=True,
        logo__isnull=False,
        description__isnull=False
    ).annotate(
        has_confirmed_booking=Exists(confirmed_bookings),
        booth_title=Subquery(booth_subquery),
        location_location=Subquery(location_subquery),
    ).filter(
        has_confirmed_booking=True
    )

    for company in companies:
        company.show_toggle = len(company.description) > 300

    return render(request, "exhibitor_list.html", {"companies": companies})
